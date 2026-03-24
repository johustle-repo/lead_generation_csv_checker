from dataclasses import dataclass
from pathlib import Path
import re
import sys
import tkinter as tk
from tkinter import filedialog, ttk
from tkinter.scrolledtext import ScrolledText

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    OPENPYXL_AVAILABLE = False


APP_TITLE = "Elmar's Lead Generation Quality Studio"
NULLABLE_FIELDS = {"Date", "Import Trades", "LinkedIn"}
OPTIONAL_COLUMNS = {"Date", "LinkedIn"}
TIMEZONE_REFERENCE_CANDIDATES = [
    "timezone_dataset.xlsx",
    "Timezone Dataset/timezone_dataset.xlsx",
]
REGION_REFERENCE_WORKBOOK_CANDIDATES = ["Regions (1).xlsx", "Regions.xlsx"]
COUNTRY_CODE_COLUMN_ALIASES = [
    "Code",
    "Country Code",
    "CountryCode",
    "Region Code",
    "ISO Code",
    "ISO2",
    "ISO 3166-1 Alpha-2",
]
EXPECTED_COLUMNS = [
    "Date",
    "Company",
    "Website",
    "First Name",
    "Email",
    "Country",
    "City",
    "Import Trades",
    "LinkedIn",
]
COLUMN_ALIASES = {
    "Date": ["Created Date", "Lead Date", "Date Added"],
    "Company": ["Company Name", "Business Name", "Organization", "Organisation"],
    "Website": ["Website URL", "Company Website", "Domain", "URL"],
    "First Name": ["Contact Person", "Contact Name", "FirstName", "Given Name"],
    "Email": ["Email Address", "Work Email", "Business Email", "Contact Email"],
    "Country": ["Country Name", "Country/Region", "Region Country", "Location"],
    "City": ["Town", "Location City", "City Name", "Capital"],
    "Import Trades": ["Import Trade", "Imports", "Trade Count", "Import Count"],
    "LinkedIn": ["LinkedIn Account", "LinkedIn URL", "LinkedIn Profile", "Linkedin"],
}
NULL_LIKE_VALUES = {"", "null", "none", "n/a", "na", "nan"}
EMAIL_REGEX = r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$"
PREVIEW_LIMIT = 250
# Central theme palettes keep ttk styles and plain Tk widgets aligned in light and dark mode.
UI_THEMES = {
    "light": {
        "root_bg": "#eee5d9",
        "card_bg": "#fffaf2",
        "dark_card_bg": "#1f2d3d",
        "soft_card_bg": "#f7efe3",
        "hero_title_fg": "#f8f4ed",
        "hero_sub_fg": "#c9d6e3",
        "hero_pill_bg": "#304155",
        "hero_pill_fg": "#e7edf4",
        "card_title_fg": "#6d4d32",
        "body_fg": "#394250",
        "card_body_fg": "#223046",
        "metric_value_fg": "#1f2d3d",
        "metric_label_fg": "#7b5b40",
        "accent_button_bg": "#d29a3a",
        "accent_button_fg": "#14202d",
        "accent_button_active": "#e3ab4a",
        "secondary_button_bg": "#ede2d2",
        "secondary_button_fg": "#223046",
        "secondary_button_active": "#f7ecdf",
        "analyze_button_bg": "#2d6a9f",
        "analyze_button_fg": "#ffffff",
        "analyze_button_active": "#3a82bf",
        "save_button_bg": "#2a7a49",
        "save_button_fg": "#ffffff",
        "save_button_active": "#35945b",
        "muted_button_bg": "#dccdb8",
        "muted_button_fg": "#3a4756",
        "muted_button_active": "#e6d8c4",
        "tree_bg": "#fffdf8",
        "tree_fg": "#1f2d3d",
        "tree_heading_bg": "#e7d7c0",
        "tree_heading_fg": "#3b2d21",
        "tree_select_bg": "#d9c2a7",
        "tree_heading_active_bg": "#dcc5ab",
        "notebook_bg": "#fffaf2",
        "tab_bg": "#e8dbc9",
        "tab_fg": "#324253",
        "tab_selected_bg": "#fffaf2",
        "tab_selected_fg": "#132434",
        "tab_active_bg": "#efe3d2",
        "summary_text_bg": "#fffdf8",
        "summary_text_fg": "#1f2d3d",
        "summary_text_select_bg": "#d9c2a7",
        "summary_text_select_fg": "#132434",
        "status_palette": {
            "neutral": {"panel": "#f5ecde", "border": "#eadcc8", "title": "#7b5b40", "detail": "#223046", "badge_bg": "#d9e2ec", "badge_fg": "#223046"},
            "success": {"panel": "#ecf8f0", "border": "#b9dfc4", "title": "#1f6f43", "detail": "#184a30", "badge_bg": "#1f6f43", "badge_fg": "#f7fff9"},
            "error": {"panel": "#fff0f2", "border": "#efc0c8", "title": "#9a2032", "detail": "#5b1824", "badge_bg": "#b52b40", "badge_fg": "#fff7f8"},
            "warning": {"panel": "#fff6eb", "border": "#edd2a8", "title": "#8b5a16", "detail": "#5c3d13", "badge_bg": "#b97718", "badge_fg": "#fff9f0"},
            "info": {"panel": "#edf5ff", "border": "#c8dced", "title": "#285b8f", "detail": "#1b3d61", "badge_bg": "#336ea8", "badge_fg": "#f6fbff"},
        },
        "tree_tags": {
            "issue_row": {"background": "#fff4b8", "foreground": "#4d3b00"},
            "duplicate_issue_row": {"background": "#f8d7da", "foreground": "#7a1f2a"},
            "schema_issue_row": {"background": "#ffe2a8", "foreground": "#5b3a00"},
            "needs_review": {"background": "#fff4b8", "foreground": "#4d3b00"},
            "duplicate_review": {"background": "#f8d7da", "foreground": "#7a1f2a"},
            "clean_row": {"background": "#fafff7", "foreground": "#1f2d3d"},
        },
    },
    "dark": {
        "root_bg": "#101722",
        "card_bg": "#17212f",
        "dark_card_bg": "#0f1722",
        "soft_card_bg": "#1b2838",
        "hero_title_fg": "#f5f8fc",
        "hero_sub_fg": "#aebfd0",
        "hero_pill_bg": "#24374b",
        "hero_pill_fg": "#e7f0fa",
        "card_title_fg": "#f0c884",
        "body_fg": "#c3d2e0",
        "card_body_fg": "#e4edf7",
        "metric_value_fg": "#f4f8fc",
        "metric_label_fg": "#a9bed1",
        "accent_button_bg": "#e0a84b",
        "accent_button_fg": "#0f1722",
        "accent_button_active": "#efba64",
        "secondary_button_bg": "#2b3a4d",
        "secondary_button_fg": "#e6eef6",
        "secondary_button_active": "#354a62",
        "analyze_button_bg": "#3d7fb5",
        "analyze_button_fg": "#ffffff",
        "analyze_button_active": "#4c95d0",
        "save_button_bg": "#2f8c5f",
        "save_button_fg": "#ffffff",
        "save_button_active": "#39a370",
        "muted_button_bg": "#425164",
        "muted_button_fg": "#d5dfeb",
        "muted_button_active": "#516379",
        "tree_bg": "#142030",
        "tree_fg": "#edf3f9",
        "tree_heading_bg": "#223448",
        "tree_heading_fg": "#f1d7a4",
        "tree_select_bg": "#38536f",
        "tree_heading_active_bg": "#2b4360",
        "notebook_bg": "#17212f",
        "tab_bg": "#223142",
        "tab_fg": "#bdd0e3",
        "tab_selected_bg": "#17212f",
        "tab_selected_fg": "#f4f8fc",
        "tab_active_bg": "#2b3d52",
        "summary_text_bg": "#142030",
        "summary_text_fg": "#edf3f9",
        "summary_text_select_bg": "#38536f",
        "summary_text_select_fg": "#ffffff",
        "status_palette": {
            "neutral": {"panel": "#1a2737", "border": "#2a3a4d", "title": "#f0c884", "detail": "#dbe7f3", "badge_bg": "#324559", "badge_fg": "#edf4fb"},
            "success": {"panel": "#162c23", "border": "#25513f", "title": "#7ad6a7", "detail": "#d9f4e6", "badge_bg": "#2a7e58", "badge_fg": "#f7fff9"},
            "error": {"panel": "#311720", "border": "#633142", "title": "#ff9dad", "detail": "#ffd8de", "badge_bg": "#a93f55", "badge_fg": "#fff7f8"},
            "warning": {"panel": "#332511", "border": "#6a4d1c", "title": "#ffcb80", "detail": "#ffe6bf", "badge_bg": "#a66a1e", "badge_fg": "#fff9f0"},
            "info": {"panel": "#16273a", "border": "#2f5378", "title": "#8fc1ff", "detail": "#daeafe", "badge_bg": "#3e72a9", "badge_fg": "#f6fbff"},
        },
        "tree_tags": {
            "issue_row": {"background": "#4f4300", "foreground": "#fff2ae"},
            "duplicate_issue_row": {"background": "#522833", "foreground": "#ffd6de"},
            "schema_issue_row": {"background": "#5b3d0f", "foreground": "#ffe2a8"},
            "needs_review": {"background": "#4f4300", "foreground": "#fff2ae"},
            "duplicate_review": {"background": "#522833", "foreground": "#ffd6de"},
            "clean_row": {"background": "#183024", "foreground": "#dcf5e6"},
        },
    },
}
POPUP_THEMES = {
    "light": {
        "success": {
            "header_bg": "#1f6f43",
            "header_fg": "#f5fff9",
            "body_bg": "#eefbf3",
            "body_fg": "#143524",
            "button_bg": "#2c8c57",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#e8ddcf",
            "secondary_button_fg": "#223046",
            "secondary_button_active_bg": "#ddd0c0",
            "secondary_button_active_fg": "#223046",
        },
        "error": {
            "header_bg": "#8d1f2d",
            "header_fg": "#fff6f7",
            "body_bg": "#fff0f2",
            "body_fg": "#4f1420",
            "button_bg": "#c63d53",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#e8ddcf",
            "secondary_button_fg": "#223046",
            "secondary_button_active_bg": "#ddd0c0",
            "secondary_button_active_fg": "#223046",
        },
        "warning": {
            "header_bg": "#8b5a16",
            "header_fg": "#fff8ec",
            "body_bg": "#fff7eb",
            "body_fg": "#5b3a0f",
            "button_bg": "#bf7a1d",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#e8ddcf",
            "secondary_button_fg": "#223046",
            "secondary_button_active_bg": "#ddd0c0",
            "secondary_button_active_fg": "#223046",
        },
        "info": {
            "header_bg": "#254e7b",
            "header_fg": "#f4f9ff",
            "body_bg": "#edf5ff",
            "body_fg": "#17324f",
            "button_bg": "#346ea8",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#e8ddcf",
            "secondary_button_fg": "#223046",
            "secondary_button_active_bg": "#ddd0c0",
            "secondary_button_active_fg": "#223046",
        },
    },
    "dark": {
        "success": {
            "header_bg": "#1a4e37",
            "header_fg": "#effff6",
            "body_bg": "#14231c",
            "body_fg": "#dcf7e8",
            "button_bg": "#2f8c5f",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#293746",
            "secondary_button_fg": "#e2ebf4",
            "secondary_button_active_bg": "#334558",
            "secondary_button_active_fg": "#ffffff",
        },
        "error": {
            "header_bg": "#671f2d",
            "header_fg": "#fff0f2",
            "body_bg": "#24141a",
            "body_fg": "#ffdbe1",
            "button_bg": "#b44458",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#293746",
            "secondary_button_fg": "#e2ebf4",
            "secondary_button_active_bg": "#334558",
            "secondary_button_active_fg": "#ffffff",
        },
        "warning": {
            "header_bg": "#6d4a16",
            "header_fg": "#fff7e8",
            "body_bg": "#241c12",
            "body_fg": "#ffe9c1",
            "button_bg": "#a97421",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#293746",
            "secondary_button_fg": "#e2ebf4",
            "secondary_button_active_bg": "#334558",
            "secondary_button_active_fg": "#ffffff",
        },
        "info": {
            "header_bg": "#24496f",
            "header_fg": "#f2f8ff",
            "body_bg": "#141f2b",
            "body_fg": "#daeafe",
            "button_bg": "#3b78b4",
            "button_fg": "#ffffff",
            "secondary_button_bg": "#293746",
            "secondary_button_fg": "#e2ebf4",
            "secondary_button_active_bg": "#334558",
            "secondary_button_active_fg": "#ffffff",
        },
    },
}
ERROR_FILL = "FFF2A8"
DUPLICATE_FILL = "F5B7B1"
HEADER_FILL = "E7D7C0"
LOGO_FILENAME = "app-logo.png"
SCANNED_DATA_FOLDER = "Scanned Data"
ERROR_DATA_FOLDER = "Data With Error"
TIMEZONE_REFERENCE_LOCATION_CACHE = None
TIMEZONE_REFERENCE_CODE_MAP_CACHE = None
REGION_REFERENCE_LOCATION_CACHE = None
REGION_REFERENCE_CODE_MAP_CACHE = None


def ensure_openpyxl():
    global Workbook, Alignment, Font, PatternFill, OPENPYXL_AVAILABLE

    if OPENPYXL_AVAILABLE:
        return True

    try:
        from openpyxl import Workbook as _Workbook
        from openpyxl.styles import Alignment as _Alignment, Font as _Font, PatternFill as _PatternFill
    except ImportError:
        return False

    Workbook = _Workbook
    Alignment = _Alignment
    Font = _Font
    PatternFill = _PatternFill
    OPENPYXL_AVAILABLE = True
    return True


def resource_path(filename):
    if hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS) / filename
    return Path(__file__).resolve().parent / filename


def normalize_country_code(value):
    if pd.isna(value):
        return ""
    return re.sub(r"[^A-Za-z0-9]+", "", str(value).strip().upper())


def first_existing_path(candidates):
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def find_timezone_reference_workbook():
    global TIMEZONE_REFERENCE_LOCATION_CACHE

    if TIMEZONE_REFERENCE_LOCATION_CACHE is not None:
        return TIMEZONE_REFERENCE_LOCATION_CACHE

    home = Path.home()
    candidates = []
    for filename in TIMEZONE_REFERENCE_CANDIDATES:
        candidates.append(resource_path(filename))
    candidates.append(home / "Desktop" / "Lead Generation" / "Dataset" / "Timezone Dataset" / "timezone_dataset.xlsx")
    candidates.append(home / "Downloads" / "timezone_dataset.xlsx")

    TIMEZONE_REFERENCE_LOCATION_CACHE = first_existing_path(candidates)
    return TIMEZONE_REFERENCE_LOCATION_CACHE


def find_region_reference_workbook():
    global REGION_REFERENCE_LOCATION_CACHE

    if REGION_REFERENCE_LOCATION_CACHE is not None:
        return REGION_REFERENCE_LOCATION_CACHE

    home = Path.home()
    candidates = []
    for filename in REGION_REFERENCE_WORKBOOK_CANDIDATES:
        candidates.append(resource_path(filename))
        candidates.append(home / "Downloads" / filename)

    REGION_REFERENCE_LOCATION_CACHE = first_existing_path(candidates)
    return REGION_REFERENCE_LOCATION_CACHE


def detect_reference_header_row(raw_df):
    for index, row in raw_df.iterrows():
        normalized_values = {
            normalize_column_name(value)
            for value in row.tolist()
            if not pd.isna(value) and str(value).strip()
        }
        if {"country", "code", "capital"}.issubset(normalized_values):
            return index
    return None


def load_timezone_reference_map():
    global TIMEZONE_REFERENCE_CODE_MAP_CACHE

    if TIMEZONE_REFERENCE_CODE_MAP_CACHE is not None:
        return TIMEZONE_REFERENCE_CODE_MAP_CACHE

    reference_path = find_timezone_reference_workbook()
    if reference_path is None:
        TIMEZONE_REFERENCE_CODE_MAP_CACHE = {}
        return TIMEZONE_REFERENCE_CODE_MAP_CACHE

    try:
        df = pd.read_excel(reference_path, dtype=str)
    except Exception:
        TIMEZONE_REFERENCE_CODE_MAP_CACHE = {}
        return TIMEZONE_REFERENCE_CODE_MAP_CACHE

    df = df.fillna("")
    df.columns = [str(column).strip() for column in df.columns]
    normalized_columns = {normalize_column_name(column): column for column in df.columns}
    original_code_column = normalized_columns.get("originalcountrycode")
    mapped_code_column = normalized_columns.get("code")
    capital_column = normalized_columns.get("capital")
    country_name_column = normalized_columns.get("country")

    if not original_code_column or not mapped_code_column or not capital_column:
        TIMEZONE_REFERENCE_CODE_MAP_CACHE = {}
        return TIMEZONE_REFERENCE_CODE_MAP_CACHE

    code_map = {}
    for _, row in df.iterrows():
        original_code = normalize_country_code(row.get(original_code_column))
        mapped_code = normalize_country_code(row.get(mapped_code_column))
        mapped_city = normalize_text(row.get(capital_column))
        country_name = normalize_text(row.get(country_name_column)) if country_name_column else ""

        if not original_code or not mapped_code or is_blank(mapped_city):
            continue

        code_map[original_code] = {
            "country_name": "" if pd.isna(country_name) else str(country_name).strip(),
            "mapped_code": mapped_code,
            "mapped_city": str(mapped_city).strip(),
        }

    TIMEZONE_REFERENCE_CODE_MAP_CACHE = code_map
    return TIMEZONE_REFERENCE_CODE_MAP_CACHE


@dataclass
class AnalysisResult:
    source_file: Path
    raw_df: pd.DataFrame
    clean_df: pd.DataFrame
    review_df: pd.DataFrame
    row_report_df: pd.DataFrame
    issue_details_df: pd.DataFrame
    null_summary_df: pd.DataFrame
    error_summary_df: pd.DataFrame
    missing_columns: list[str]

    @property
    def total_rows(self):
        return len(self.raw_df)

    @property
    def clean_rows(self):
        return len(self.clean_df)

    @property
    def issue_rows(self):
        return len(self.row_report_df)

    @property
    def total_issues(self):
        return len(self.issue_details_df)

    @property
    def has_issues(self):
        return self.total_issues > 0 or bool(self.missing_columns)

    @property
    def issue_row_numbers(self):
        if self.row_report_df.empty or "Row Number" not in self.row_report_df.columns:
            return []
        row_numbers = []
        for value in self.row_report_df["Row Number"].tolist():
            try:
                row_numbers.append(int(value))
            except (TypeError, ValueError):
                continue
        return sorted(set(row_numbers))


def is_blank(value):
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip().lower() in NULL_LIKE_VALUES:
        return True
    return False


def normalize_text(value):
    if pd.isna(value):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in NULL_LIKE_VALUES:
            return pd.NA
        return stripped
    return value


def is_effectively_blank_row(row):
    return all(is_blank(value) for value in row.tolist())


def detect_email_column(columns):
    resolved_columns = resolve_expected_columns(columns)
    return resolved_columns.get("Email")


def normalize_column_name(column_name):
    return re.sub(r"[^a-z0-9]+", "", str(column_name).strip().lower())


def resolve_expected_columns(columns):
    normalized_lookup = {normalize_column_name(column): column for column in columns}
    resolved_columns = {}

    for expected_column in EXPECTED_COLUMNS:
        candidates = [expected_column] + COLUMN_ALIASES.get(expected_column, [])
        normalized_candidates = sorted(
            {normalize_column_name(candidate) for candidate in candidates},
            key=len,
            reverse=True,
        )

        for normalized_candidate in normalized_candidates:
            actual_column = normalized_lookup.get(normalized_candidate)
            if actual_column:
                resolved_columns[expected_column] = actual_column
                break

    return resolved_columns


def detect_country_code_column(columns):
    normalized_lookup = {normalize_column_name(column): column for column in columns}
    for candidate in COUNTRY_CODE_COLUMN_ALIASES:
        actual_column = normalized_lookup.get(normalize_column_name(candidate))
        if actual_column:
            return actual_column
    return None


def values_look_like_country_codes(series, reference_codes):
    sample_values = []
    for value in series.tolist():
        normalized = normalize_country_code(value)
        if normalized:
            sample_values.append(normalized)
        if len(sample_values) >= 25:
            break

    if not sample_values:
        return False

    matched = sum(1 for value in sample_values if value in reference_codes)
    minimum_required = max(1, (len(sample_values) * 3 + 4) // 5)
    return matched >= minimum_required


def apply_reference_country_city_mapping(df, resolved_expected_columns):
    code_map = load_timezone_reference_map()
    if not code_map:
        return df, resolved_expected_columns, []

    code_column = detect_country_code_column(df.columns)
    if code_column is None:
        country_column = resolved_expected_columns.get("Country")
        if country_column and values_look_like_country_codes(df[country_column], set(code_map)):
            code_column = country_column

    if code_column is None:
        return df, resolved_expected_columns, []

    country_column = resolved_expected_columns.get("Country")
    if country_column is None:
        df["Country"] = pd.NA
        country_column = "Country"
        resolved_expected_columns["Country"] = "Country"

    city_column = resolved_expected_columns.get("City")
    if city_column is None:
        df["City"] = pd.NA
        city_column = "City"
        resolved_expected_columns["City"] = "City"

    corrections = []
    for index, row in df.iterrows():
        normalized_code = normalize_country_code(row.get(code_column))
        if not normalized_code:
            continue

        reference_entry = code_map.get(normalized_code)
        if not reference_entry:
            continue

        # Keep the original code field synced with the mapped export code from the timezone reference.
        corrected_code = reference_entry["mapped_code"]
        if code_column != country_column:
            original_code = row.get(code_column)
            original_code_text = "" if pd.isna(original_code) else str(original_code).strip()
            if original_code_text != corrected_code:
                corrections.append(
                    {
                        "row_index": index,
                        "column": code_column,
                        "original_value": original_code_text,
                        "corrected_value": corrected_code,
                        "issue_type": "Correction",
                        "problem": f'Code updated to "{corrected_code}" based on timezone dataset',
                    }
                )
            df.at[index, code_column] = corrected_code

        original_country = row.get(country_column)
        original_country_text = "" if pd.isna(original_country) else str(original_country).strip()
        corrected_country = corrected_code
        if original_country_text != corrected_country:
            corrections.append(
                {
                    "row_index": index,
                    "column": country_column,
                    "original_value": original_country_text,
                    "corrected_value": corrected_country,
                    "issue_type": "Correction",
                    "problem": f'Country updated to "{corrected_country}" based on timezone dataset',
                }
            )
        df.at[index, country_column] = corrected_country

        original_city = row.get(city_column)
        original_city_text = "" if pd.isna(original_city) else str(original_city).strip()
        corrected_city = reference_entry["mapped_city"]
        if original_city_text != corrected_city:
            corrections.append(
                {
                    "row_index": index,
                    "column": city_column,
                    "original_value": original_city_text,
                    "corrected_value": corrected_city,
                    "issue_type": "Correction",
                    "problem": f'City updated to "{corrected_city}" based on timezone dataset',
                }
            )
        df.at[index, city_column] = corrected_city

    return df, resolved_expected_columns, corrections


def canonical_export_column(column_name, resolved_expected_columns):
    for expected_column, actual_column in resolved_expected_columns.items():
        if actual_column == column_name:
            return expected_column
    return column_name


def format_row_list(row_numbers):
    if not row_numbers:
        return "None"
    if len(row_numbers) <= 12:
        return ", ".join(str(row) for row in row_numbers)
    preview = ", ".join(str(row) for row in row_numbers[:12])
    return f"{preview}, ... (+{len(row_numbers) - 12} more)"


def build_issue_row_map(issue_details_df):
    issue_row_map = {}
    if issue_details_df.empty:
        return issue_row_map

    row_level_issues = issue_details_df[issue_details_df["Row Number"] != "Schema"]
    for column, group in row_level_issues.groupby("Column"):
        row_numbers = []
        for value in group["Row Number"].tolist():
            try:
                row_numbers.append(int(value))
            except (TypeError, ValueError):
                continue
        issue_row_map[column] = sorted(set(row_numbers))
    return issue_row_map


def build_duplicate_email_map(issue_details_df):
    duplicate_email_map = {}
    if issue_details_df.empty or "Issue Type" not in issue_details_df.columns:
        return duplicate_email_map

    duplicate_rows = issue_details_df[
        (issue_details_df["Issue Type"] == "Duplicate")
        & (issue_details_df["Column"] == "Email")
    ]
    if duplicate_rows.empty:
        return duplicate_email_map

    for issue in duplicate_rows.itertuples(index=False):
        row_number = issue[0]
        email_value = "" if pd.isna(issue[4]) else str(issue[4]).strip().lower()
        if not email_value:
            continue
        duplicate_email_map.setdefault(email_value, []).append(int(row_number))

    return {
        email: sorted(set(row_numbers))
        for email, row_numbers in duplicate_email_map.items()
    }


def build_review_dataframe(source_df, issue_details_df, missing_columns):
    review_df = source_df.copy().astype(object)
    resolved_expected_columns = resolve_expected_columns(review_df.columns)

    for column in missing_columns:
        review_df[column] = "[MISSING COLUMN]"

    for expected_column, actual_column in resolved_expected_columns.items():
        if expected_column not in review_df.columns and actual_column in review_df.columns:
            review_df[expected_column] = review_df[actual_column]

    ordered_columns = []
    for column in EXPECTED_COLUMNS:
        if column in review_df.columns:
            ordered_columns.append(column)

    alias_source_columns = {
        actual_column
        for expected_column, actual_column in resolved_expected_columns.items()
        if actual_column != expected_column
    }
    extra_columns = [
        column for column in review_df.columns if column not in ordered_columns and column not in alias_source_columns
    ]
    review_df = review_df[ordered_columns + extra_columns]

    if "Row Number" in review_df.columns:
        review_df["Source Row Number"] = review_df["Row Number"]
        review_df = review_df.drop(columns=["Row Number"])
    if "Row Status" in review_df.columns:
        review_df = review_df.drop(columns=["Row Status"])
    if "Error Columns" in review_df.columns:
        review_df = review_df.drop(columns=["Error Columns"])
    if "Issues Found" in review_df.columns:
        review_df = review_df.drop(columns=["Issues Found"])

    review_df.insert(0, "Row Number", [index + 2 for index in range(len(review_df))])
    review_df.insert(1, "Row Status", "Clean")
    review_df["Error Columns"] = ""
    review_df["Issues Found"] = ""

    if issue_details_df.empty:
        return review_df

    row_level_issues = issue_details_df[issue_details_df["Row Number"] != "Schema"].copy()
    if row_level_issues.empty:
        return review_df

    row_to_messages = {}
    row_to_columns = {}
    for issue in row_level_issues.itertuples(index=False):
        row_number = int(issue[0])
        column = issue[1]
        problem = issue[3]

        row_to_messages.setdefault(row_number, []).append(f"{column}: {problem}")
        row_to_columns.setdefault(row_number, set()).add(column)

        row_index = row_number - 2
        if 0 <= row_index < len(review_df) and column in review_df.columns:
            current_value = review_df.at[row_index, column]
            if pd.isna(current_value) or str(current_value).strip() == "":
                review_df.at[row_index, column] = f"[ERROR] {problem}"
            elif not str(current_value).startswith("[ERROR]"):
                review_df.at[row_index, column] = f"[ERROR] {current_value}"

    for row_number, messages in row_to_messages.items():
        row_index = row_number - 2
        review_df.at[row_index, "Row Status"] = "Needs Review"
        review_df.at[row_index, "Issues Found"] = " | ".join(messages)
        review_df.at[row_index, "Error Columns"] = ", ".join(sorted(row_to_columns[row_number]))

    return review_df


def build_excel_export_dataframe(review_df):
    export_df = review_df.copy()
    helper_columns = [
        "Row Number",
        "Row Status",
        "Error Columns",
        "Issues Found",
        "Source Row Number",
    ]
    export_df = export_df.drop(columns=[column for column in helper_columns if column in export_df.columns])

    for column in export_df.columns:
        export_df[column] = export_df[column].apply(clean_export_value)

    return export_df


def clean_export_value(value):
    if pd.isna(value):
        return ""
    text = str(value)
    if text == "[MISSING COLUMN]":
        return ""
    if text == "[ERROR] Null or blank value":
        return ""
    if text.startswith("[ERROR] "):
        return text.replace("[ERROR] ", "", 1)
    return text


def is_valid_email(email):
    if is_blank(email):
        return False
    return re.match(EMAIL_REGEX, str(email).strip()) is not None


def analyze_csv(file_path):
    source_file = Path(file_path)

    try:
        df = pd.read_csv(source_file, dtype=str)
    except Exception as exc:
        raise ValueError(f"Unable to read CSV:\n{exc}") from exc

    if df.empty:
        raise ValueError("The selected CSV file is empty.")

    df.columns = [str(col).strip() for col in df.columns]
    for col in df.columns:
        df[col] = df[col].apply(normalize_text)

    # Ignore spreadsheet padding rows so trailing empty lines are not exported as highlighted errors.
    df = df.loc[~df.apply(is_effectively_blank_row, axis=1)].reset_index(drop=True)
    if df.empty:
        raise ValueError("The selected CSV contains only blank rows.")

    resolved_expected_columns = resolve_expected_columns(df.columns)
    df, resolved_expected_columns, field_corrections = apply_reference_country_city_mapping(df, resolved_expected_columns)
    missing_columns = [
        column
        for column in EXPECTED_COLUMNS
        if column not in OPTIONAL_COLUMNS and column not in resolved_expected_columns
    ]
    email_column = detect_email_column(df.columns)
    required_fields = [
        col
        for col in df.columns
        if canonical_export_column(col, resolved_expected_columns) not in NULLABLE_FIELDS
    ]

    email_duplicates = set()
    if email_column:
        normalized_emails = (
            df[email_column]
            .dropna()
            .astype(str)
            .str.strip()
            .str.lower()
        )
        duplicate_mask = normalized_emails.duplicated(keep=False)
        email_duplicates = set(normalized_emails[duplicate_mask].tolist())

    clean_rows = []
    row_report_rows = []
    issue_detail_rows = []

    for missing_column in missing_columns:
        issue_detail_rows.append(
            {
                "Row Number": "Schema",
                "Column": missing_column,
                "Issue Type": "Missing Column",
                "Problem": "Expected column is missing from the file",
                "Current Value": "",
            }
        )

    correction_map = {}
    for correction in field_corrections:
        correction_map.setdefault(correction["row_index"], []).append(correction)

    for idx, row in df.iterrows():
        row_number = idx + 2
        row_issue_text = []

        for correction in correction_map.get(idx, []):
            display_column = canonical_export_column(correction["column"], resolved_expected_columns)
            row_issue_text.append(f'{display_column}: {correction["problem"]}')
            issue_detail_rows.append(
                {
                    "Row Number": row_number,
                    "Column": correction["column"],
                    "Issue Type": correction["issue_type"],
                    "Problem": correction["problem"],
                    "Current Value": correction["original_value"],
                }
            )

        for field in required_fields:
            if is_blank(row.get(field)):
                issue_message = "Null or blank value"
                row_issue_text.append(f"{field}: {issue_message}")
                issue_detail_rows.append(
                    {
                        "Row Number": row_number,
                        "Column": field,
                        "Issue Type": "Null",
                        "Problem": issue_message,
                        "Current Value": "",
                    }
                )

        if email_column and not is_blank(row.get(email_column)):
            email_value = str(row[email_column]).strip()
            if not is_valid_email(email_value):
                issue_message = "Invalid email format"
                row_issue_text.append(f"{email_column}: {issue_message}")
                issue_detail_rows.append(
                    {
                        "Row Number": row_number,
                        "Column": email_column,
                        "Issue Type": "Format",
                        "Problem": issue_message,
                        "Current Value": email_value,
                    }
                )
            elif email_value.lower() in email_duplicates:
                issue_message = "Duplicate email detected"
                row_issue_text.append(f"{email_column}: {issue_message}")
                issue_detail_rows.append(
                    {
                        "Row Number": row_number,
                        "Column": email_column,
                        "Issue Type": "Duplicate",
                        "Problem": issue_message,
                        "Current Value": email_value,
                    }
                )

        clean_row_data = row.to_dict()
        row_data = row.to_dict()
        row_data["Row Number"] = row_number

        if row_issue_text:
            row_data["Issues Found"] = " | ".join(row_issue_text)
            row_report_rows.append(row_data)
        else:
            clean_rows.append(clean_row_data)

    clean_df = pd.DataFrame(clean_rows)
    row_report_df = pd.DataFrame(row_report_rows)
    issue_details_df = pd.DataFrame(issue_detail_rows)
    review_df = build_review_dataframe(df, issue_details_df, missing_columns)

    if issue_details_df.empty:
        null_summary_df = pd.DataFrame(columns=["Column", "Count"])
        error_summary_df = pd.DataFrame(columns=["Column", "Count"])
    else:
        null_summary_df = (
            issue_details_df[issue_details_df["Issue Type"] == "Null"]
            .groupby("Column")
            .size()
            .reset_index(name="Count")
            .sort_values(["Count", "Column"], ascending=[False, True])
        )
        error_summary_df = (
            issue_details_df[issue_details_df["Issue Type"] != "Null"]
            .groupby("Column")
            .size()
            .reset_index(name="Count")
            .sort_values(["Count", "Column"], ascending=[False, True])
        )

    return AnalysisResult(
        source_file=source_file,
        raw_df=df,
        clean_df=clean_df,
        review_df=review_df,
        row_report_df=row_report_df,
        issue_details_df=issue_details_df,
        null_summary_df=null_summary_df,
        error_summary_df=error_summary_df,
        missing_columns=missing_columns,
    )


def default_clean_path(source_file):
    return default_output_directory(source_file, SCANNED_DATA_FOLDER) / f"{source_file.stem}_cleaned_usable.csv"


def default_review_path(source_file):
    return default_output_directory(source_file, ERROR_DATA_FOLDER) / f"{source_file.stem}_reviewed_with_errors.xlsx"


def default_issue_path(source_file):
    return default_output_directory(source_file, ERROR_DATA_FOLDER) / f"{source_file.stem}_issue_details.csv"


def default_output_directory(source_file, folder_name):
    output_directory = Path(source_file).resolve().parent / folder_name
    output_directory.mkdir(parents=True, exist_ok=True)
    return output_directory


def make_unique_path(path):
    path = Path(path)
    if not path.exists():
        return path

    counter = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{counter}{path.suffix}")
        if not candidate.exists():
            return candidate
        counter += 1


def save_dataframe(df, title, suggested_path):
    save_path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=".csv",
        initialfile=suggested_path.name,
        initialdir=str(suggested_path.parent),
        filetypes=[("CSV Files", "*.csv"), ("All files", "*.*")],
    )
    if not save_path:
        return None

    # Excel on Windows reads accented characters reliably when the CSV includes a UTF-8 BOM.
    df.to_csv(save_path, index=False, encoding="utf-8-sig")
    return Path(save_path)


def write_csv_file(df, save_path):
    save_path = Path(save_path)
    save_path.parent.mkdir(parents=True, exist_ok=True)
    # Excel on Windows reads accented characters reliably when the CSV includes a UTF-8 BOM.
    df.to_csv(save_path, index=False, encoding="utf-8-sig")
    return save_path


def write_review_workbook(review_df, issue_details_df, missing_columns, save_path):
    if not ensure_openpyxl():
        raise RuntimeError(
            "Excel export requires the 'openpyxl' package. Install it with: pip install openpyxl"
        )

    save_path = Path(save_path)
    save_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reviewed Export"
    export_df = build_excel_export_dataframe(review_df)
    resolved_expected_columns = resolve_expected_columns(review_df.columns)

    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    error_fill = PatternFill(fill_type="solid", fgColor=ERROR_FILL)
    duplicate_fill = PatternFill(fill_type="solid", fgColor=DUPLICATE_FILL)
    header_font = Font(bold=True, color="223046")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    columns = list(export_df.columns)
    column_positions = {column: index + 1 for index, column in enumerate(columns)}

    for column_index, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    issue_cells = {}
    if "Row Number" in issue_details_df.columns:
        row_level_issues = issue_details_df[issue_details_df["Row Number"] != "Schema"]
    else:
        row_level_issues = pd.DataFrame(columns=["Row Number", "Column"])
    for issue in row_level_issues.itertuples(index=False):
        row_number = int(issue[0])
        column_name = canonical_export_column(issue[1], resolved_expected_columns)
        if column_name in column_positions:
            excel_row = row_number
            excel_column = column_positions[column_name]
            fill_kind = "duplicate" if str(issue[2]) == "Duplicate" else "error"
            existing_kind = issue_cells.get((excel_row, excel_column))
            if existing_kind != "duplicate":
                issue_cells[(excel_row, excel_column)] = fill_kind

    for column_name in missing_columns:
        if column_name in column_positions:
            worksheet.cell(row=1, column=column_positions[column_name]).fill = error_fill
            for data_row in range(2, len(review_df) + 2):
                issue_cells[(data_row, column_positions[column_name])] = "error"

    for row_index, row in enumerate(export_df.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            column_name = columns[column_index - 1]
            if column_name == "Date":
                cell = worksheet.cell(row=row_index, column=column_index, value="=TODAY()")
                cell.number_format = "mm/dd/yyyy"
            else:
                display_value = "" if pd.isna(value) else str(value)
                cell = worksheet.cell(row=row_index, column=column_index, value=display_value)
            cell.alignment = wrap_alignment

            cell_fill_kind = issue_cells.get((row_index, column_index))
            if cell_fill_kind == "duplicate":
                cell.fill = duplicate_fill
            elif cell_fill_kind == "error":
                cell.fill = error_fill

    for column_index, column_name in enumerate(columns, start=1):
        max_width = len(str(column_name))
        for row_index in range(2, len(export_df) + 2):
            cell_value = worksheet.cell(row=row_index, column=column_index).value
            if cell_value is None:
                continue
            max_width = max(max_width, len(str(cell_value)))
        worksheet.column_dimensions[worksheet.cell(row=1, column=column_index).column_letter].width = min(max_width + 2, 40)

    worksheet.freeze_panes = "A2"
    try:
        workbook.save(save_path)
    except Exception as exc:
        raise RuntimeError(f"Could not save the Excel file:\n{exc}") from exc
    return Path(save_path)


def save_review_workbook(review_df, issue_details_df, missing_columns, title, suggested_path):
    save_path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=".xlsx",
        initialfile=suggested_path.name,
        initialdir=str(suggested_path.parent),
        filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
    )
    if not save_path:
        return None

    return write_review_workbook(
        review_df,
        issue_details_df,
        missing_columns,
        save_path,
    )


class PremiumCSVCheckerApp:
    def __init__(self, root):
        self.root = root
        self.current_theme_name = "light"
        self.style = ttk.Style()
        self.root.title(APP_TITLE)
        self.root.geometry("1240x820")
        self.root.minsize(1040, 720)
        self.root.configure(bg=self.get_theme()["root_bg"])

        self.selected_file = None
        self.analysis_result = None
        self.icon_logo_image = None
        self.hero_logo_image = None
        self.current_status_tone = "neutral"
        self.current_status_badge_text = "Waiting For Upload"
        self.current_status_title = "Status Overview"

        self.file_var = tk.StringVar(value="No file uploaded yet")
        self.status_var = tk.StringVar(value="Upload a CSV file to begin the analysis.")
        self.summary_var = tk.StringVar(
            value="The analyzer will inspect null values, invalid emails, duplicates, and row-level issues."
        )
        self.schema_var = tk.StringVar(
            value=f"Expected columns: {', '.join(EXPECTED_COLUMNS)}"
        )

        self.total_rows_var = tk.StringVar(value="0")
        self.clean_rows_var = tk.StringVar(value="0")
        self.issue_rows_var = tk.StringVar(value="0")
        self.total_issues_var = tk.StringVar(value="0")
        self.status_badge_var = tk.StringVar(value="Waiting For Upload")
        self.export_button_var = tk.StringVar(value="Save Export CSV")
        self.theme_button_var = tk.StringVar()

        self.configure_styles()
        self.load_brand_assets()
        self.build_ui()
        self.update_status_banner(
            "neutral",
            "Waiting For Upload",
            "Status Overview",
            "Upload a CSV file to begin the analysis.",
        )
        self.apply_theme()

    def get_theme(self):
        return UI_THEMES[self.current_theme_name]

    def update_theme_button_label(self):
        if self.current_theme_name == "dark":
            self.theme_button_var.set("Switch to Light Mode")
        else:
            self.theme_button_var.set("Switch to Dark Mode")

    def apply_theme(self):
        theme = self.get_theme()
        self.root.configure(bg=theme["root_bg"])
        self.configure_styles()
        self.update_theme_button_label()

        # ttk widgets pick up colors from styles; these direct updates cover the remaining tk widgets.
        if hasattr(self, "hero_logo_label") and self.hero_logo_label is not None:
            self.hero_logo_label.configure(bg=theme["dark_card_bg"])
        if hasattr(self, "hero_stats"):
            self.hero_stats.configure(bg=theme["hero_pill_bg"], fg=theme["hero_pill_fg"])
        if hasattr(self, "summary_text"):
            self.summary_text.configure(
                bg=theme["summary_text_bg"],
                fg=theme["summary_text_fg"],
                insertbackground=theme["summary_text_fg"],
                selectbackground=theme["summary_text_select_bg"],
                selectforeground=theme["summary_text_select_fg"],
            )
        if hasattr(self, "issue_tree_widget"):
            for tag_name, colors in theme["tree_tags"].items():
                self.issue_tree_widget.tag_configure(tag_name, **colors)
        if hasattr(self, "clean_tree"):
            for tag_name, colors in theme["tree_tags"].items():
                self.clean_tree.tag_configure(tag_name, **colors)

            self.update_status_banner(
            self.current_status_tone,
            self.current_status_badge_text,
            self.current_status_title,
        )

    def toggle_theme(self):
        # Flip between the two palette sets and repaint the UI in place.
        self.current_theme_name = "dark" if self.current_theme_name == "light" else "light"
        self.apply_theme()

    def load_brand_assets(self):
        logo_path = resource_path(LOGO_FILENAME)
        if not logo_path.exists():
            return

        try:
            self.icon_logo_image = tk.PhotoImage(file=str(logo_path))
            self.root.iconphoto(True, self.icon_logo_image)

            scale_factor = max(
                1,
                max(
                    (self.icon_logo_image.width() + 139) // 140,
                    (self.icon_logo_image.height() + 139) // 140,
                ),
            )
            self.hero_logo_image = self.icon_logo_image.subsample(scale_factor, scale_factor)
        except tk.TclError:
            self.icon_logo_image = None
            self.hero_logo_image = None

    def configure_styles(self):
        theme = self.get_theme()
        style = self.style
        if "clam" in style.theme_names():
            style.theme_use("clam")

        style.configure("App.TFrame", background=theme["root_bg"])
        style.configure("Card.TFrame", background=theme["card_bg"])
        style.configure("DarkCard.TFrame", background=theme["dark_card_bg"])
        style.configure("SoftCard.TFrame", background=theme["soft_card_bg"])
        style.configure(
            "Hero.TLabel",
            background=theme["dark_card_bg"],
            foreground=theme["hero_title_fg"],
            font=("Segoe UI Semibold", 24),
        )
        style.configure(
            "HeroSub.TLabel",
            background=theme["dark_card_bg"],
            foreground=theme["hero_sub_fg"],
            font=("Segoe UI", 10),
        )
        style.configure(
            "CardTitle.TLabel",
            background=theme["card_bg"],
            foreground=theme["card_title_fg"],
            font=("Segoe UI Semibold", 11),
        )
        style.configure(
            "Body.TLabel",
            background=theme["root_bg"],
            foreground=theme["body_fg"],
            font=("Segoe UI", 10),
        )
        style.configure(
            "CardBody.TLabel",
            background=theme["card_bg"],
            foreground=theme["card_body_fg"],
            font=("Segoe UI", 10),
        )
        style.configure(
            "MetricValue.TLabel",
            background=theme["card_bg"],
            foreground=theme["metric_value_fg"],
            font=("Segoe UI Semibold", 26),
        )
        style.configure(
            "MetricLabel.TLabel",
            background=theme["card_bg"],
            foreground=theme["metric_label_fg"],
            font=("Segoe UI", 9),
        )
        style.configure(
            "Accent.TButton",
            font=("Segoe UI Semibold", 10),
            padding=(16, 10),
            background=theme["accent_button_bg"],
            foreground=theme["accent_button_fg"],
            borderwidth=0,
        )
        style.configure(
            "Secondary.TButton",
            font=("Segoe UI", 10),
            padding=(14, 9),
            background=theme["secondary_button_bg"],
            foreground=theme["secondary_button_fg"],
            borderwidth=0,
        )
        style.configure(
            "Theme.TButton",
            font=("Segoe UI Semibold", 10),
            padding=(14, 9),
            background=theme["secondary_button_bg"],
            foreground=theme["secondary_button_fg"],
            borderwidth=0,
        )
        style.configure(
            "Analyze.TButton",
            font=("Segoe UI Semibold", 10),
            padding=(16, 10),
            background=theme["analyze_button_bg"],
            foreground=theme["analyze_button_fg"],
            borderwidth=0,
        )
        style.configure(
            "Save.TButton",
            font=("Segoe UI Semibold", 10),
            padding=(14, 9),
            background=theme["save_button_bg"],
            foreground=theme["save_button_fg"],
            borderwidth=0,
        )
        style.configure(
            "Muted.TButton",
            font=("Segoe UI", 10),
            padding=(14, 9),
            background=theme["muted_button_bg"],
            foreground=theme["muted_button_fg"],
            borderwidth=0,
        )
        style.map("Accent.TButton", background=[("active", theme["accent_button_active"])])
        style.map("Analyze.TButton", background=[("active", theme["analyze_button_active"])])
        style.map("Save.TButton", background=[("active", theme["save_button_active"])])
        style.map("Secondary.TButton", background=[("active", theme["secondary_button_active"])])
        style.map("Theme.TButton", background=[("active", theme["secondary_button_active"])])
        style.map("Muted.TButton", background=[("active", theme["muted_button_active"])])
        style.configure(
            "Premium.Treeview",
            background=theme["tree_bg"],
            foreground=theme["tree_fg"],
            fieldbackground=theme["tree_bg"],
            rowheight=28,
            borderwidth=0,
            font=("Segoe UI", 9),
        )
        style.configure(
            "Premium.Treeview.Heading",
            background=theme["tree_heading_bg"],
            foreground=theme["tree_heading_fg"],
            font=("Segoe UI Semibold", 9),
            relief="flat",
        )
        style.map("Premium.Treeview", background=[("selected", theme["tree_select_bg"])])
        style.map(
            "Premium.Treeview.Heading",
            background=[("active", theme["tree_heading_active_bg"])],
        )
        style.configure(
            "Premium.TNotebook",
            background=theme["notebook_bg"],
            borderwidth=0,
        )
        style.configure(
            "Premium.TNotebook.Tab",
            background=theme["tab_bg"],
            foreground=theme["tab_fg"],
            padding=(18, 10),
            font=("Segoe UI Semibold", 9),
        )
        style.map(
            "Premium.TNotebook.Tab",
            background=[("selected", theme["tab_selected_bg"]), ("active", theme["tab_active_bg"])],
            foreground=[("selected", theme["tab_selected_fg"])],
        )
        style.configure(
            "Vertical.TScrollbar",
            background=theme["card_bg"],
            troughcolor=theme["soft_card_bg"],
            borderwidth=0,
            arrowcolor=theme["card_body_fg"],
        )
        style.configure(
            "Horizontal.TScrollbar",
            background=theme["card_bg"],
            troughcolor=theme["soft_card_bg"],
            borderwidth=0,
            arrowcolor=theme["card_body_fg"],
        )

    def build_ui(self):
        theme = self.get_theme()
        wrapper = ttk.Frame(self.root, padding=18, style="App.TFrame")
        wrapper.pack(fill="both", expand=True)
        wrapper.columnconfigure(0, weight=1)
        wrapper.rowconfigure(3, weight=1)

        hero = ttk.Frame(wrapper, padding=22, style="DarkCard.TFrame")
        hero.grid(row=0, column=0, sticky="ew")
        hero.columnconfigure(1, weight=1)

        if self.hero_logo_image is not None:
            self.hero_logo_label = tk.Label(
                hero,
                image=self.hero_logo_image,
                bg=theme["dark_card_bg"],
                bd=0,
                highlightthickness=0,
            )
            self.hero_logo_label.grid(row=0, column=0, rowspan=3, sticky="w", padx=(0, 20))
        else:
            self.hero_logo_label = None

        ttk.Label(hero, text="Lead List Quality Studio", style="Hero.TLabel").grid(
            row=0, column=1, sticky="w"
        )
        ttk.Label(
            hero,
            text="Elmar's premium checker for lead-generation CSVs, with schema checks, row-level diagnostics, and clean export control.",
            style="HeroSub.TLabel",
        ).grid(row=1, column=1, sticky="w", pady=(8, 0))
        self.hero_stats = tk.Label(
            hero,
            text="Built for faster CSV review before Reply.io upload",
            bg=theme["hero_pill_bg"],
            fg=theme["hero_pill_fg"],
            font=("Segoe UI Semibold", 9),
            padx=12,
            pady=6,
        )
        self.hero_stats.grid(row=2, column=1, sticky="w", pady=(16, 0))

        controls = ttk.Frame(wrapper, padding=18, style="Card.TFrame")
        controls.grid(row=1, column=0, sticky="ew", pady=(16, 14))
        controls.columnconfigure(1, weight=1)
        controls.columnconfigure(2, weight=0)

        ttk.Label(controls, text="Uploaded File", style="CardTitle.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(
            controls,
            textvariable=self.file_var,
            style="CardBody.TLabel",
            wraplength=720,
        ).grid(row=0, column=1, sticky="w", padx=(14, 0))
        self.status_badge = tk.Label(
            controls,
            textvariable=self.status_badge_var,
            bg=theme["status_palette"]["neutral"]["badge_bg"],
            fg=theme["status_palette"]["neutral"]["badge_fg"],
            font=("Segoe UI Semibold", 9),
            padx=12,
            pady=7,
        )
        self.status_badge.grid(row=0, column=2, sticky="e")

        button_row = ttk.Frame(controls, style="Card.TFrame")
        button_row.grid(row=1, column=0, columnspan=3, sticky="w", pady=(18, 0))

        ttk.Button(
            button_row,
            text="Upload File",
            command=self.upload_file,
            style="Accent.TButton",
        ).grid(row=0, column=0, padx=(0, 10))

        ttk.Button(
            button_row,
            text="Analyze",
            command=self.run_analysis,
            style="Analyze.TButton",
        ).grid(row=0, column=1, padx=(0, 10))

        self.save_clean_button = ttk.Button(
            button_row,
            textvariable=self.export_button_var,
            command=self.save_clean_file,
            style="Save.TButton",
        )
        self.save_clean_button.grid(row=0, column=2, padx=(0, 10))
        self.save_clean_button.state(["disabled"])

        self.save_issue_button = ttk.Button(
            button_row,
            text="Save Issue Report",
            command=self.save_issue_report,
            style="Muted.TButton",
        )
        self.save_issue_button.grid(row=0, column=3)
        self.save_issue_button.state(["disabled"])

        ttk.Button(
            button_row,
            textvariable=self.theme_button_var,
            command=self.toggle_theme,
            style="Theme.TButton",
        ).grid(row=0, column=4, padx=(10, 0))

        status_panel = tk.Frame(
            controls,
            bg=theme["status_palette"]["neutral"]["panel"],
            highlightthickness=1,
            highlightbackground=theme["status_palette"]["neutral"]["border"],
        )
        status_panel.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(18, 0))
        status_panel.grid_columnconfigure(0, weight=1)
        self.status_panel = status_panel

        self.status_title_label = tk.Label(
            status_panel,
            text="Status Overview",
            bg=theme["status_palette"]["neutral"]["panel"],
            fg=theme["status_palette"]["neutral"]["title"],
            font=("Segoe UI Semibold", 10),
            anchor="w",
            padx=14,
            pady=8,
        )
        self.status_title_label.grid(row=0, column=0, sticky="ew")
        self.status_detail_label = tk.Label(
            status_panel,
            textvariable=self.status_var,
            bg=theme["status_palette"]["neutral"]["panel"],
            fg=theme["status_palette"]["neutral"]["detail"],
            font=("Segoe UI", 10),
            anchor="w",
            justify="left",
            padx=14,
            pady=2,
        )
        self.status_detail_label.grid(row=1, column=0, sticky="ew")
        self.schema_detail_label = tk.Label(
            status_panel,
            textvariable=self.schema_var,
            bg=theme["status_palette"]["neutral"]["panel"],
            fg=theme["status_palette"]["neutral"]["detail"],
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
            padx=14,
            pady=8,
        )
        self.schema_detail_label.grid(row=2, column=0, sticky="ew")

        metrics = ttk.Frame(wrapper, style="App.TFrame")
        metrics.grid(row=2, column=0, sticky="ew", pady=(2, 0))
        for idx in range(4):
            metrics.columnconfigure(idx, weight=1)

        self.create_metric_card(metrics, 0, "Total Rows", self.total_rows_var)
        self.create_metric_card(metrics, 1, "Clean Rows", self.clean_rows_var)
        self.create_metric_card(metrics, 2, "Rows With Issues", self.issue_rows_var)
        self.create_metric_card(metrics, 3, "Total Issues", self.total_issues_var)

        notebook_card = ttk.Frame(wrapper, padding=14, style="Card.TFrame")
        notebook_card.grid(row=3, column=0, sticky="nsew", pady=(16, 0))
        notebook_card.columnconfigure(0, weight=1)
        notebook_card.rowconfigure(0, weight=1)

        self.notebook = ttk.Notebook(notebook_card, style="Premium.TNotebook")
        self.notebook.grid(row=0, column=0, sticky="nsew")

        self.summary_frame = ttk.Frame(self.notebook, style="Card.TFrame")
        self.issue_frame = ttk.Frame(self.notebook, style="Card.TFrame")
        self.clean_frame = ttk.Frame(self.notebook, style="Card.TFrame")

        self.notebook.add(self.summary_frame, text="Summary")
        self.notebook.add(self.issue_frame, text="Issue Details")
        self.notebook.add(self.clean_frame, text="Export Preview")

        self.build_summary_tab()
        self.build_issue_tab()
        self.build_clean_tab()

    def create_metric_card(self, parent, column, label, variable):
        card = ttk.Frame(parent, padding=16, style="Card.TFrame")
        card.grid(row=0, column=column, sticky="nsew", padx=(0 if column == 0 else 10, 0))
        ttk.Label(card, text=variable.get(), textvariable=variable, style="MetricValue.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(card, text=label, style="MetricLabel.TLabel").grid(
            row=1, column=0, sticky="w", pady=(8, 0)
        )

    def update_status_banner(self, tone, badge_text, status_title, detail_text=None):
        self.current_status_tone = tone
        self.current_status_badge_text = badge_text
        self.current_status_title = status_title
        colors = self.get_theme()["status_palette"][tone]
        self.status_badge_var.set(badge_text)
        self.status_badge.configure(bg=colors["badge_bg"], fg=colors["badge_fg"])
        self.status_panel.configure(bg=colors["panel"], highlightbackground=colors["border"])
        self.status_title_label.configure(bg=colors["panel"], fg=colors["title"], text=status_title)
        self.status_detail_label.configure(bg=colors["panel"], fg=colors["detail"])
        self.schema_detail_label.configure(bg=colors["panel"], fg=colors["detail"])
        if detail_text is not None:
            self.status_var.set(detail_text)

    def show_custom_popup(
        self,
        title,
        message,
        tone="info",
        ask=False,
        primary_text=None,
        secondary_text="Not Now",
    ):
        theme = POPUP_THEMES[self.current_theme_name][tone]
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.configure(bg=theme["body_bg"])
        popup.resizable(False, False)
        popup.transient(self.root)
        if self.icon_logo_image is not None:
            popup.iconphoto(True, self.icon_logo_image)

        popup_width = 540
        popup_height = 360 if ask else 320
        popup.withdraw()
        popup.geometry(f"{popup_width}x{popup_height}")

        header = tk.Frame(popup, bg=theme["header_bg"], height=72)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(
            header,
            text=title,
            bg=theme["header_bg"],
            fg=theme["header_fg"],
            font=("Segoe UI Semibold", 16),
            anchor="w",
            padx=20,
        ).pack(fill="both", expand=True)

        body = tk.Frame(popup, bg=theme["body_bg"], padx=22, pady=18)
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(0, weight=1)

        message_box = ScrolledText(
            body,
            wrap="word",
            font=("Segoe UI", 10),
            bg=theme["body_bg"],
            fg=theme["body_fg"],
            relief="flat",
            borderwidth=0,
            highlightthickness=0,
            padx=2,
            pady=2,
        )
        message_box.grid(row=0, column=0, sticky="nsew")
        message_box.insert("1.0", message)
        message_box.configure(state="disabled")

        button_row = tk.Frame(body, bg=theme["body_bg"])
        button_row.grid(row=1, column=0, sticky="ew", pady=(16, 0))

        result = {"value": None}

        def close_with(value):
            result["value"] = value
            popup.destroy()

        self.center_popup(popup, popup_width, popup_height)
        popup.deiconify()
        popup.grab_set()
        popup.focus_force()

        if ask:
            tk.Button(
                button_row,
                text=primary_text or "Continue",
                command=lambda: close_with(True),
                bg=theme["button_bg"],
                fg=theme["button_fg"],
                activebackground=theme["button_bg"],
                activeforeground=theme["button_fg"],
                relief="flat",
                font=("Segoe UI Semibold", 10),
                padx=16,
                pady=8,
                cursor="hand2",
            ).pack(side="left")
            tk.Button(
                button_row,
                text=secondary_text,
                command=lambda: close_with(False),
                bg=theme["secondary_button_bg"],
                fg=theme["secondary_button_fg"],
                activebackground=theme["secondary_button_active_bg"],
                activeforeground=theme["secondary_button_active_fg"],
                relief="flat",
                font=("Segoe UI", 10),
                padx=16,
                pady=8,
                cursor="hand2",
            ).pack(side="left", padx=(10, 0))
        else:
            tk.Button(
                button_row,
                text="Close",
                command=lambda: close_with(True),
                bg=theme["button_bg"],
                fg=theme["button_fg"],
                activebackground=theme["button_bg"],
                activeforeground=theme["button_fg"],
                relief="flat",
                font=("Segoe UI Semibold", 10),
                padx=16,
                pady=8,
                cursor="hand2",
            ).pack(side="left")

        popup.protocol("WM_DELETE_WINDOW", lambda: close_with(False if ask else True))
        self.root.wait_window(popup)
        return result["value"]

    def center_popup(self, popup, width, height):
        self.root.update_idletasks()
        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()

        if root_width <= 1 or root_height <= 1:
            screen_width = popup.winfo_screenwidth()
            screen_height = popup.winfo_screenheight()
            x = max((screen_width - width) // 2, 0)
            y = max((screen_height - height) // 2, 0)
        else:
            x = max(root_x + (root_width - width) // 2, 0)
            y = max(root_y + (root_height - height) // 2, 0)

        popup.geometry(f"{width}x{height}+{x}+{y}")

    def build_summary_tab(self):
        self.summary_frame.columnconfigure(0, weight=1)
        self.summary_frame.rowconfigure(1, weight=1)

        summary_header = ttk.Frame(self.summary_frame, padding=18, style="Card.TFrame")
        summary_header.grid(row=0, column=0, sticky="ew")
        summary_header.columnconfigure(0, weight=1)

        ttk.Label(summary_header, text="Analyzer Summary", style="CardTitle.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(
            summary_header,
            textvariable=self.summary_var,
            style="CardBody.TLabel",
            wraplength=1040,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=(8, 0))

        self.summary_text = ScrolledText(
            self.summary_frame,
            wrap="word",
            font=("Consolas", 10),
            bg="#fffdf8",
            fg="#1f2d3d",
            relief="flat",
            padx=10,
            pady=10,
        )
        self.summary_text.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        self.set_text_widget(self.summary_text, "Upload a CSV file and click Analyze to populate the dashboard.")

    def build_issue_tab(self):
        self.issue_frame.columnconfigure(0, weight=1)
        self.issue_frame.rowconfigure(1, weight=1)

        ttk.Label(
            self.issue_frame,
            text="Rows, columns, and exact issue messages appear here after analysis.",
            style="CardBody.TLabel",
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 10))

        self.issue_tree_container, self.issue_tree_widget = self.create_treeview(
            self.issue_frame,
            ["Row Number", "Column", "Issue Type", "Problem", "Current Value"],
        )
        self.issue_tree_widget.tag_configure(
            "issue_row",
            background="#fff4b8",
            foreground="#4d3b00",
        )
        self.issue_tree_widget.tag_configure(
            "duplicate_issue_row",
            background="#f8d7da",
            foreground="#7a1f2a",
        )
        self.issue_tree_widget.tag_configure(
            "schema_issue_row",
            background="#ffe2a8",
            foreground="#5b3a00",
        )
        self.issue_tree_container.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))

    def build_clean_tab(self):
        self.clean_frame.columnconfigure(0, weight=1)
        self.clean_frame.rowconfigure(1, weight=1)

        ttk.Label(
            self.clean_frame,
            text="The downloadable export is previewed here. Clean files show the clean export; files with issues show the reviewed export with error markers.",
            style="CardBody.TLabel",
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 10))

        self.clean_tree = ttk.Treeview(self.clean_frame, show="headings", style="Premium.Treeview")
        self.clean_tree.tag_configure(
            "needs_review",
            background="#fff4b8",
            foreground="#4d3b00",
        )
        self.clean_tree.tag_configure(
            "duplicate_review",
            background="#f8d7da",
            foreground="#7a1f2a",
        )
        self.clean_tree.tag_configure(
            "clean_row",
            background="#fafff7",
            foreground="#1f2d3d",
        )
        clean_scroll_y = ttk.Scrollbar(self.clean_frame, orient="vertical", command=self.clean_tree.yview)
        clean_scroll_x = ttk.Scrollbar(self.clean_frame, orient="horizontal", command=self.clean_tree.xview)
        self.clean_tree.configure(yscrollcommand=clean_scroll_y.set, xscrollcommand=clean_scroll_x.set)

        self.clean_tree.grid(row=1, column=0, sticky="nsew", padx=(18, 0), pady=(0, 18))
        clean_scroll_y.grid(row=1, column=1, sticky="ns", pady=(0, 18), padx=(0, 18))
        clean_scroll_x.grid(row=2, column=0, sticky="ew", padx=(18, 0))

    def create_treeview(self, parent, columns):
        container = ttk.Frame(parent, style="Card.TFrame")
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        tree = ttk.Treeview(container, columns=columns, show="headings", style="Premium.Treeview")
        scroll_y = ttk.Scrollbar(container, orient="vertical", command=tree.yview)
        scroll_x = ttk.Scrollbar(container, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        for column in columns:
            tree.heading(column, text=column)
            width = 140 if column != "Problem" and column != "Current Value" else 260
            tree.column(column, width=width, anchor="w")

        tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")
        return container, tree

    def upload_file(self):
        file_path = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV Files", "*.csv"), ("All files", "*.*")],
        )
        if not file_path:
            return

        self.selected_file = Path(file_path)
        self.analysis_result = None
        self.file_var.set(str(self.selected_file))
        self.status_var.set("File uploaded. Click Analyze to inspect the data quality.")
        self.summary_var.set(
            "The file is loaded. Run the analyzer to see null counts, invalid data, and cleaned-row availability."
        )
        self.schema_var.set(f"Expected columns: {', '.join(EXPECTED_COLUMNS)}")
        self.total_rows_var.set("0")
        self.clean_rows_var.set("0")
        self.issue_rows_var.set("0")
        self.total_issues_var.set("0")
        self.export_button_var.set("Save Export CSV")
        self.save_clean_button.state(["disabled"])
        self.save_issue_button.state(["disabled"])
        self.issue_tree_widget.delete(*self.issue_tree_widget.get_children())
        self.reset_clean_preview()
        self.set_text_widget(self.summary_text, "File uploaded successfully. Click Analyze to generate the report.")
        self.update_status_banner(
            "info",
            "File Loaded",
            "Upload Complete",
            "The file is loaded. Click Analyze to inspect the data quality and export options.",
        )

    def run_analysis(self):
        if not self.selected_file:
            self.show_custom_popup(
                "Upload Required",
                "Please upload a CSV file first.",
                tone="info",
            )
            return

        try:
            result = analyze_csv(self.selected_file)
        except ValueError as exc:
            self.update_status_banner(
                "error",
                "Read Error",
                "Analysis Error",
                "The analyzer could not read the selected file.",
            )
            self.summary_var.set(str(exc))
            self.set_text_widget(self.summary_text, str(exc))
            self.show_custom_popup("Analysis Error", str(exc), tone="error")
            return

        self.analysis_result = result
        self.update_dashboard(result)
        self.refresh_issue_table(result.issue_details_df)
        self.refresh_clean_preview(result.review_df if result.has_issues else result.clean_df)

        if result.has_issues and not result.review_df.empty:
            self.export_button_var.set("Download Highlighted Error File")
            self.save_clean_button.state(["!disabled"])
        elif result.clean_rows > 0:
            self.export_button_var.set("Save Cleaned CSV")
            self.save_clean_button.state(["!disabled"])
        else:
            self.export_button_var.set("Save Export CSV")
            self.save_clean_button.state(["disabled"])

        if result.total_issues > 0:
            self.save_issue_button.state(["!disabled"])
        else:
            self.save_issue_button.state(["disabled"])

        if result.has_issues:
            self.update_status_banner(
                "error",
                "Issues Found",
                "Action Required",
                "The file contains errors or missing columns. You can still download a reviewed CSV with the missing columns added and the error rows marked.",
            )
            popup_message = self.build_issue_popup(result)
            should_download = self.show_custom_popup(
                "Issues Found",
                popup_message + "\n\nDo you want to download the highlighted error file?",
                tone="error",
                ask=True,
                primary_text="Download Highlighted Error File",
                secondary_text="Not Now",
            )
            if should_download:
                self.save_clean_file()
        else:
            self.update_status_banner(
                "success",
                "Upload Ready",
                "Ready For Export",
                "No null values or format issues were found. This file is clean and ready for Reply.io export.",
            )
            should_save = self.show_custom_popup(
                "Ready To Save",
                "No null values or format errors were found.\n\nDo you want to save the cleaned CSV now?",
                tone="success",
                ask=True,
                primary_text="Save Cleaned CSV",
                secondary_text="Not Now",
            )
            if should_save:
                self.save_clean_file()

    def update_dashboard(self, result):
        self.total_rows_var.set(str(result.total_rows))
        self.clean_rows_var.set(str(result.clean_rows))
        self.issue_rows_var.set(str(result.issue_rows))
        self.total_issues_var.set(str(result.total_issues))
        self.summary_var.set(self.build_summary_text(result))
        self.set_text_widget(self.summary_text, self.build_long_report(result))

    def build_summary_text(self, result):
        if result.has_issues:
            row_text = format_row_list(result.issue_row_numbers)
            missing_text = (
                f" Missing columns: {', '.join(result.missing_columns)}."
                if result.missing_columns
                else ""
            )
            top_problem = (
                f"Error rows: {row_text}.{missing_text} You can download a highlighted Excel file with the missing columns added and the problematic cells marked in yellow."
            )
        else:
            top_problem = "Everything passed. The CSV is ready for a clean export."

        return (
            f"Rows checked: {result.total_rows} | Clean rows: {result.clean_rows} | "
            f"Rows with issues: {result.issue_rows} | Total issue entries: {result.total_issues}. "
            f"{top_problem}"
        )

    def build_long_report(self, result):
        duplicate_email_map = build_duplicate_email_map(result.issue_details_df)
        lines = [
            f"File: {result.source_file}",
            f"Total Rows: {result.total_rows}",
            f"Clean Rows: {result.clean_rows}",
            f"Rows With Issues: {result.issue_rows}",
            f"Total Issue Entries: {result.total_issues}",
            f"Rows with Errors: {format_row_list(result.issue_row_numbers)}",
            "",
            "Nullable Fields:",
            ", ".join(sorted(NULLABLE_FIELDS)),
        ]

        if result.missing_columns:
            lines.extend(["", "Missing Columns:"])
            for column in result.missing_columns:
                lines.append(f"- {column}")
        else:
            lines.extend(["", "Missing Columns:", "No expected columns are missing."])

        if result.null_summary_df.empty:
            lines.extend(["", "Null Summary:", "No null or blank values found in required fields."])
        else:
            lines.extend(["", "Null Summary:"])
            for _, row in result.null_summary_df.iterrows():
                lines.append(f"- {row['Column']}: {row['Count']} null/blank issue(s)")

        if result.error_summary_df.empty:
            lines.extend(["", "Other Errors:", "No email format or duplicate-email issues found."])
        else:
            lines.extend(["", "Other Errors:"])
            for _, row in result.error_summary_df.iterrows():
                lines.append(f"- {row['Column']}: {row['Count']} non-null issue(s)")

        if duplicate_email_map:
            lines.extend(["", "Duplicate Emails:"])
            for email, row_numbers in sorted(duplicate_email_map.items()):
                lines.append(f"- {email}: rows {format_row_list(row_numbers)}")
        else:
            lines.extend(["", "Duplicate Emails:", "No duplicate emails found."])

        issue_row_map = build_issue_row_map(result.issue_details_df)
        if issue_row_map:
            lines.extend(["", "Issue Rows By Column:"])
            for column, row_numbers in sorted(issue_row_map.items()):
                lines.append(f"- {column}: rows {format_row_list(row_numbers)}")

        if result.issue_details_df.empty:
            lines.extend(["", "Result:", "This file is clean and ready to save."])
        else:
            lines.extend(["", "Result:", "The file has issues, but you can export a reviewed CSV with missing columns added and error cells marked."])

        return "\n".join(lines)

    def build_issue_popup(self, result):
        issue_row_map = build_issue_row_map(result.issue_details_df)
        duplicate_email_map = build_duplicate_email_map(result.issue_details_df)
        lines = [
            f"Rows with issues: {result.issue_rows}",
            f"Total issue entries: {result.total_issues}",
            f"Row numbers with errors: {format_row_list(result.issue_row_numbers)}",
            "",
            "Missing columns:",
        ]

        if result.missing_columns:
            for column in result.missing_columns:
                lines.append(f"- {column}")
        else:
            lines.append("- None")

        lines.extend([
            "",
            "Columns with null or blank values:",
        ])

        if result.null_summary_df.empty:
            lines.append("- None")
        else:
            for _, row in result.null_summary_df.iterrows():
                row_numbers = issue_row_map.get(row["Column"], [])
                lines.append(
                    f"- {row['Column']}: {row['Count']} | rows {format_row_list(row_numbers)}"
                )

        lines.append("")
        lines.append("Columns with other errors:")
        if result.error_summary_df.empty:
            lines.append("- None")
        else:
            for _, row in result.error_summary_df.iterrows():
                row_numbers = issue_row_map.get(row["Column"], [])
                lines.append(
                    f"- {row['Column']}: {row['Count']} | rows {format_row_list(row_numbers)}"
                )

        lines.extend(["", "Duplicate emails:"])
        if duplicate_email_map:
            for email, row_numbers in sorted(duplicate_email_map.items()):
                lines.append(f"- {email} | rows {format_row_list(row_numbers)}")
        else:
            lines.append("- None")

        return "\n".join(lines)

    def refresh_issue_table(self, issue_df):
        self.populate_tree(self.issue_tree_widget, issue_df)

    def refresh_clean_preview(self, clean_df):
        columns = list(clean_df.columns)
        self.clean_tree.delete(*self.clean_tree.get_children())

        if not columns:
            self.clean_tree["columns"] = ("Status",)
            self.clean_tree.heading("Status", text="Status")
            self.clean_tree.column("Status", width=260, anchor="w")
            self.clean_tree.insert("", "end", values=("No clean rows available yet.",))
            return

        preview_df = clean_df.head(PREVIEW_LIMIT)
        self.clean_tree["columns"] = columns
        for column in columns:
            self.clean_tree.heading(column, text=column)
            self.clean_tree.column(column, width=160, anchor="w")

        status_index = columns.index("Row Status") if "Row Status" in columns else None
        for row in preview_df.itertuples(index=False):
            values = ["" if pd.isna(value) else str(value) for value in row]
            tags = ()
            issues_index = columns.index("Issues Found") if "Issues Found" in columns else None
            issues_text = values[issues_index] if issues_index is not None else ""
            if "Duplicate email detected" in issues_text:
                tags = ("duplicate_review",)
            elif status_index is not None and values[status_index] == "Needs Review":
                tags = ("needs_review",)
            elif status_index is not None:
                tags = ("clean_row",)
            self.clean_tree.insert("", "end", values=values, tags=tags)

    def populate_tree(self, tree, df):
        tree.delete(*tree.get_children())
        for row in df.itertuples(index=False):
            values = ["" if pd.isna(value) else str(value) for value in row]
            tags = ()
            if tree is self.issue_tree_widget:
                if values and values[0] == "Schema":
                    tags = ("schema_issue_row",)
                elif len(values) > 2 and values[2] == "Duplicate":
                    tags = ("duplicate_issue_row",)
                else:
                    tags = ("issue_row",)
            tree.insert("", "end", values=values, tags=tags)

    def reset_clean_preview(self):
        self.clean_tree.delete(*self.clean_tree.get_children())
        self.clean_tree["columns"] = ("Status",)
        self.clean_tree.heading("Status", text="Status")
        self.clean_tree.column("Status", width=260, anchor="w")
        self.clean_tree.insert("", "end", values=("Analyze a file to preview the cleaned rows.",))

    def save_clean_file(self):
        if not self.analysis_result:
            self.show_custom_popup(
                "Analyze Required",
                "Run the analyzer before saving.",
                tone="info",
            )
            return
        if self.analysis_result.has_issues:
            if self.analysis_result.review_df.empty:
                self.show_custom_popup(
                    "Nothing To Save",
                    "There is no highlighted error file available to export.",
                    tone="warning",
                )
                return

            try:
                save_path = write_review_workbook(
                    self.analysis_result.review_df,
                    self.analysis_result.issue_details_df,
                    self.analysis_result.missing_columns,
                    make_unique_path(default_review_path(self.analysis_result.source_file)),
                )
            except RuntimeError as exc:
                self.show_custom_popup(
                    "Excel Export Unavailable",
                    str(exc),
                    tone="warning",
                )
                return
            except Exception as exc:
                self.show_custom_popup(
                    "Save Failed",
                    f"The highlighted error file could not be saved.\n\n{exc}",
                    tone="error",
                )
                return

            if save_path:
                self.update_status_banner(
                    "warning",
                    "Highlighted File Saved",
                    "Review Export Complete",
                    f"Highlighted error file saved to: {save_path}",
                )
                self.show_custom_popup(
                    "Saved",
                    "The highlighted Excel file was saved with missing columns added and row-level errors filled in yellow.\n\n"
                    f"Location:\n{save_path}",
                    tone="warning",
                )
            return

        if self.analysis_result.clean_df.empty:
            self.show_custom_popup(
                "Nothing To Save",
                "There are no clean rows available to save.",
                tone="warning",
            )
            return

        try:
            save_path = write_csv_file(
                self.analysis_result.clean_df,
                make_unique_path(default_clean_path(self.analysis_result.source_file)),
            )
        except Exception as exc:
            self.show_custom_popup(
                "Save Failed",
                f"The cleaned CSV could not be saved.\n\n{exc}",
                tone="error",
            )
            return
        if save_path:
            self.update_status_banner(
                "success",
                "Clean File Saved",
                "Export Complete",
                f"Cleaned CSV saved to: {save_path}",
            )
            self.show_custom_popup("Saved", f"Cleaned CSV saved to:\n{save_path}", tone="success")

    def save_issue_report(self):
        if not self.analysis_result:
            self.show_custom_popup(
                "Analyze Required",
                "Run the analyzer before saving.",
                tone="info",
            )
            return
        if self.analysis_result.issue_details_df.empty:
            self.show_custom_popup(
                "No Issue Report",
                "There are no issue details to export.",
                tone="warning",
            )
            return

        try:
            save_path = write_csv_file(
                self.analysis_result.issue_details_df,
                make_unique_path(default_issue_path(self.analysis_result.source_file)),
            )
        except Exception as exc:
            self.show_custom_popup(
                "Save Failed",
                f"The issue report could not be saved.\n\n{exc}",
                tone="error",
            )
            return
        if save_path:
            self.update_status_banner(
                "warning",
                "Issue Report Saved",
                "Report Exported",
                f"Issue report saved to: {save_path}",
            )
            self.show_custom_popup("Saved", f"Issue report saved to:\n{save_path}", tone="warning")

    def set_text_widget(self, widget, text):
        widget.configure(state="normal")
        widget.delete("1.0", "end")
        widget.insert("1.0", text)
        widget.configure(state="disabled")


def main():
    root = tk.Tk()
    app = PremiumCSVCheckerApp(root)
    app.reset_clean_preview()
    root.mainloop()


if __name__ == "__main__":
    main()
